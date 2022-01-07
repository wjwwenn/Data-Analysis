import openpyxl
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator #copy formula

wb = openpyxl.load_workbook('test.xlsx')
sheet = wb.active

sheet = wb.worksheets[0]
max_row_num = sheet.max_row

print("All sheet names = {} " .format(wb.sheetnames))

sheet['A5']='=AVERAGE(A2:A4)'
sheet['B5'] = Translator('=AVERAGE(A2:A4)', origin="A5").translate_formula("B5")

wb.save('test.xlsx')
